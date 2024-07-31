from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook
import pandas as pd
import datetime
import smtplib
import json
import os
from flask import Flask, request, render_template, redirect, url_for, send_from_directory, flash, session, jsonify
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xls', 'xlsx'}
app.secret_key = 'supersecretkey'  # Needed for flash messages
upcoming_dates=[]

# User data for demonstration purposes
users = {
    "admin": generate_password_hash("adminpassword"),
    "user1": generate_password_hash("user1password")
}

# Function to check if user is admin
def is_admin():
    return session.get('username') == 'admin'

def load_config(config_filename):
    with open(config_filename, 'r') as config_file:
        config_dict = json.load(config_file)
    return config_dict

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def is_weekend(date):
    return date.weekday() >= 5  # Saturday and Sunday are 5 and 6

def future_dates(config):
    start_date = datetime.date.today()
    end_date=config.get('END_DATE')
    server_limit = config.get('SERVER_LIMIT') or 10
    if not end_date:
        end_date = start_date + datetime.timedelta(days=45)
    else:
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
    plan_weekends = config.get('PLAN_WEEKENDS') == 'yes' or "yes"
    # Generate the date range
    schedule_dates = []
    current_date = start_date
    while current_date <= end_date:
        if plan_weekends and current_date.weekday() in [5,6]:  # Weekdays are 0-4
            schedule_dates.append(current_date.strftime('%Y-%m-%d'))
        elif (not plan_weekends) and current_date.weekday() in [0,1,2,3,4]:
            schedule_dates.append(current_date.strftime('%Y-%m-%d'))
        current_date += datetime.timedelta(days=1)
    return schedule_dates

def schedule_maintenance(data, config):
    schedule_dates=future_dates(config)
    server_limit = config.get('SERVER_LIMIT') or 10
    #schedule = {date: [] for date in schedule_dates}
    # Check if there are enough dates to schedule all servers
    total_slots = len(schedule_dates) * server_limit
    total_servers = sum(len(records) for records in data.values())
    if total_servers > int(total_slots):
        raise ValueError("Not enough scheduling slots to fit all servers.")
    
    # Distribute servers across the available dates
    server_count = 0
    for sheet, records in data.items():
        flat_schedule_dict={}
        flat_schedule_dict[sheet]=[]
        for record in records:
            date_index = server_count // int(server_limit)
            if date_index >= len(schedule_dates):
                raise ValueError("Index out of range, more servers than available slots.")
            schedule_date = schedule_dates[date_index]
            upcoming_date = schedule_date
            record['enddate'] = upcoming_date
            upcoming_dates.append(upcoming_date)
            flat_schedule_dict[sheet].append(record)
            server_count += 1
    
    return flat_schedule_dict, upcoming_dates

@app.route('/login', methods=['GET', 'POST'])
def login():
    current_year = datetime.datetime.now().year
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username in users and check_password_hash(users[username], password):
            session['username'] = username
            flash('Login successful!', 'success')
            return redirect(url_for('timeslots'))
        else:
            flash('Invalid credentials', 'danger')
    return render_template('login.html', current_year=current_year)

@app.route('/logout')
def logout():
    session.pop('username', None)
    flash('Logged out successfully!', 'success')
    return redirect(url_for('login'))

@app.route('/download_template')
def download_template():
    # Replace 'static' with the directory where your template file is stored
    return send_from_directory(directory='static', path='template.xlsx', as_attachment=True)

@app.route('/')
def index():
    return redirect('timeslots')

@app.route('/uploads')
def uploads():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'username' not in session or not is_admin():
        flash('Access denied', 'danger')
        return redirect(url_for('login'))
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    
    if file and allowed_file(file.filename):
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)
        wb = load_workbook(filepath)
        sheet_names = wb.sheetnames
        config_filename=file.filename.split('.')[0]+'.config'
        config_dict={}
        config_dict["END_DATE"]=request.form.get("start")
        config_dict["PLAN_WEEKENDS"]=request.form.get("plan-weekends", "no")
        config_dict["SERVER_LIMIT"]=request.form["server-limit"]
        config_dict["MAINT_LISTG"]=request.form["server-listing"]
        config_dict["FILENAME"]=file.filename
        with open(config_filename, 'w') as jsonfile:
            json.dump(config_dict, jsonfile)
        data = {}
        for sheet in sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet)
            data[sheet] = df.to_dict('records')
        session['filename'] = file.filename
        return render_template('customize.html', data=data, filename=file.filename, sheet_names=sheet_names)
    flash('Invalid file type', 'danger')
    return redirect(request.url)

@app.route('/customize', methods=['POST'])
def customize():
    filename = request.form['filename']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    #custom_data = []
    wb = load_workbook(filepath)
    sheet_names = wb.sheetnames
    
    for sheet in sheet_names:
        ws = wb[sheet]
        custom_servers = request.form.getlist(f'custom_server_{sheet}')
        custom_emails = request.form.getlist(f'custom_email_{sheet}')
        custom_paths = request.form.getlist(f'custom_path_{sheet}')
        custom_arrays = request.form.getlist(f'custom_array_{sheet}')
        custom_storages = request.form.getlist(f'custom_storage_{sheet}')
        custom_enddates = request.form.getlist(f'custom_enddate_{sheet}')
        custom_snow_changes = request.form.getlist(f'custom_snow_change_{sheet}')
        custom_maintenance_names = request.form.getlist(f'custom_maintenance_name_{sheet}')
        
        for i, (custom_server, custom_email, custom_path, custom_array, custom_storage, custom_enddate, custom_snow_change, custom_maintenance_name) in enumerate(zip(custom_servers, custom_emails, custom_paths, custom_arrays, custom_storages, custom_enddates, custom_snow_changes, custom_maintenance_names), start=2):
            ws.cell(row=i, column=1, value=custom_server)
            ws.cell(row=i, column=2, value=custom_email)
            ws.cell(row=i, column=3, value=custom_path)
            ws.cell(row=i, column=4, value=custom_array)
            ws.cell(row=i, column=5, value=custom_storage)
            ws.cell(row=i, column=6, value=custom_enddate)
            ws.cell(row=i, column=7, value=custom_snow_change)
            ws.cell(row=i, column=8, value=custom_maintenance_name)
        
        wb.save(filepath)
    
    flash('Custom slots saved successfully', 'success')
    return redirect(url_for('view_timeslots', filename=filename))


@app.route('/timeslots')
def timeslots():
    if 'username' not in session:
        flash('Please login to view this page', 'warning')
        return redirect(url_for('login'))
    filenames = os.listdir(app.config['UPLOAD_FOLDER'])
    filenames = [f for f in filenames if allowed_file(f)]  # Filter only allowed files
    return render_template('timeslot.html', filenames=filenames)

@app.route('/view_timeslots')
def view_timeslots():
    if 'username' not in session:
        flash('Please login to view this page', 'warning')
        return redirect(url_for('login'))
    filename = request.args.get('filename')
    session['filename']=filename
    if not filename:
        return redirect(url_for('timeslots'))
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.isfile(filepath):
        flash('File not found', 'danger')
        return redirect(url_for('timeslots'))
    config_filename=filename.split('.')[0]+'.config'
    if os.path.isfile(config_filename):
        #with open(config_filename, 'r') as config_file:
        config_dict=load_config(config_filename)
            #config_dict = json.load(config_file)
    wb = load_workbook(filepath)
    sheet_names = wb.sheetnames
    data = {}
    for sheet in sheet_names:
        df = pd.read_excel(filepath, sheet_name=sheet)
        df['url'] = df['email'].apply(lambda email: url_for('server_details', email=email,filename=filename))
        data[sheet] = df.to_dict('records')
    new_schedule, upcoming_dates = schedule_maintenance(data, config_dict)
    upcoming_dates = list(set(upcoming_dates))
    session['upcoming_dates']=upcoming_dates
    return render_template('timeslots.html', data=new_schedule, filename=filename, sheet_names=sheet_names, upcoming_dates=upcoming_dates)


@app.route('/update_slots', methods=['POST'])
def update_slots():
    filename = request.form['filename']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if not os.path.exists(filepath):
        flash('File does not exist', 'error')
        return redirect(url_for('view_timeslots', filename=filename))

    wb = load_workbook(filepath)

    # Retrieve the sheets from the workbook
    sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # Iterate over slots based on form data
        i = 0
        while f'servers_{sheet_name}_{i}' in request.form:
            custom_server = request.form.get(f'servers_{sheet_name}_{i}')
            custom_enddate = request.form.get(f'enddate_{sheet_name}_{i}_dropdown', request.form.get(f'enddate_{sheet_name}_{i}'))
            acknowledgment = request.form.get(f'acknowledgment_{sheet_name}_{i}')
            notification = request.form.get(f'notification_{sheet_name}_{i}')

            # Debug print statements
            print(f'Updating row {i+2} in sheet {sheet_name}: Server: {custom_server}, End Date: {custom_enddate}, Notification: {notification}, Acknowledgment: {acknowledgment}')

            ws.cell(row=i + 2, column=1, value=custom_server)
            ws.cell(row=i + 2, column=6, value=custom_enddate)
            ws.cell(row=i + 2, column=9, value='Yes' if notification else 'No')  # Assuming column 9 for notification
            ws.cell(row=i + 2, column=10, value='Yes' if acknowledgment else 'No')  # Assuming column 10 for acknowledgment

            i += 1

    wb.save(filepath)
    flash('All slots updated successfully', 'success')
    return redirect(url_for('view_timeslots', filename=filename))


@app.route('/update_slot', methods=['POST'])
def update_slot():
    filename = request.form['filename']
    slot_index = int(request.form['slot_index'])
    email = request.form['email']
    custom_enddate = request.form.get('enddate_dropdown', request.form.get('enddate'))
    acknowledgment = request.form.get(f'acknowledgment_{slot_index}')
    notification = request.form.get(f'notification_{slot_index}')

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb = load_workbook(filepath)
    ws = wb['Sheet1']  # Adjust sheet name as needed

    # Debug print statements
    print(f'Updating row {slot_index + 2} in sheet: End Date: {custom_enddate}, Notification: {notification}, Acknowledgment: {acknowledgment}')

    # Update the cells (adjust column indices as needed)
    ws.cell(row=slot_index + 2, column=6, value=custom_enddate)
    ws.cell(row=slot_index + 2, column=9, value='Yes' if notification else 'No')
    ws.cell(row=slot_index + 2, column=10, value='Yes' if acknowledgment else 'No')

    wb.save(filepath)
    flash('Slot updated successfully', 'success')
    return redirect(url_for('view_timeslots', filename=filename))

@app.route('/update_slot_ajax', methods=['POST'])
def update_slot_ajax():
    filename = request.form['filename']
    server_name = request.form['server_name']
    email = request.form['email']
    custom_enddate = request.form.get('enddate_dropdown', request.form.get('enddate'))
    acknowledgment = request.form.get('acknowledgment') == 'true'
    notification = request.form.get('notification') == 'true'
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb = load_workbook(filepath)
    ws = wb['Sheet1']  # Adjust sheet name as needed

    # Find the row that matches the server name
    row_to_update = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value == server_name:
            row_to_update = row[0].row
            break

    if row_to_update is None:
        flash('Server not found', 'error')
        return jsonify({'status': 'error', 'message': 'Server not found'})

    # Debug print statements
    print(f'Updating row {row_to_update} in sheet: End Date: {custom_enddate}, Notification: {notification}, Acknowledgment: {acknowledgment}')

    # Update the cells (adjust column indices as needed)
    ws.cell(row=row_to_update, column=6, value=custom_enddate)
    ws.cell(row=row_to_update, column=9, value='Yes' if notification else 'No')
    ws.cell(row=row_to_update, column=10, value='Yes')

    wb.save(filepath)
    flash('Slot updated successfully', 'success')
    return jsonify({'status': 'success'})


@app.route('/send_reminder', methods=['POST'])
def send_reminder():
    filename = request.form['filename']
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb = load_workbook(filepath)
    sheet_names = wb.sheetnames
    emails_sent = []

    for sheet in sheet_names:
        df = pd.read_excel(filepath, sheet_name=sheet)
        for _, row in df.iterrows():
            if row['acknowledgment'] == 'yes':  # Only send to acknowledged slots
                server = row['servers']
                email = row['email']
                
                maintenance_name = row['maintenance_name']
                enddate = row['enddate']
                message = MIMEMultipart()
                message['From'] = 'your_email@example.com'
                message['To'] = email
                message['Subject'] = 'Maintenance Reminder'
                body = f'Dear User,\n\nThis is a reminder for the maintenance of server {server}.\nMaintenance Name: {maintenance_name}\nEnd Date: {enddate}\n\nThank you.'
                message.attach(MIMEText(body, 'plain'))
                try:
                    with smtplib.SMTP('smtp.example.com', 587) as server:
                        server.starttls()
                        server.login('your_email@example.com', 'your_password')
                        server.sendmail('your_email@example.com', email, message.as_string())
                    emails_sent.append(email)
                except Exception as e:
                    print(f'Failed to send email to {email}: {str(e)}')
                    flash(f'Failed to send email to {email}', 'danger')
    
    flash(f'Successfully sent emails to: {", ".join(emails_sent)}', 'success')
    return redirect(url_for('timeslots', filename=filename))

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/server_details/<path:email>', methods=['GET', 'POST'])
def server_details(email):
    maintname = request.args.get('maintname')
    filename=maintname+".xlsx"
    if not filename:
        flash('Filename is missing', 'danger')
        return redirect(url_for('timeslots'))

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.isfile(filepath):
        flash('File not found', 'danger')
        return redirect(url_for('timeslots'))

    wb = load_workbook(filepath)
    sheet_names = wb.sheetnames
    server_data = []

    for sheet in sheet_names:
        df = pd.read_excel(filepath, sheet_name=sheet)
        for _, row in df.iterrows():
            if row['email'] == email:
                server_data.append(row.to_dict())

    if not server_data:
        flash('No server data found for this email', 'warning')
        return redirect(url_for('timeslots'))

    unique_url = url_for('server_details', email=email, _external=True)
    
    config_filename=maintname+".config"
    config=load_config(config_filename)
    upcoming_maintenance_dates=future_dates(config)
    
    return render_template('server_details.html', server_data=server_data, filename=filename, unique_url=unique_url, upcoming_maintenance_dates=upcoming_maintenance_dates)



if __name__ == '__main__':
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    app.run(debug=True)
